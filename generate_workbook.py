
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import datetime

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────────────────────────────────────
# Crypto sheet  – Deep Navy / Gold theme
C_DARK      = "0D1B2A"   # deep navy bg
C_MID       = "1B2E45"   # section header bg
C_ACCENT    = "D4AF37"   # gold
C_LIGHT     = "E8F1FA"   # light blue tint
C_WHITE     = "FFFFFF"
C_GREEN     = "28A745"
C_RED       = "DC3545"
C_YELLOW    = "FFC107"

# Eqaaz sheet – Emerald / White / Light-Gold Islamic theme
E_DARK      = "064635"   # deep emerald
E_MID       = "0A6645"   # section header
E_ACCENT    = "C8A951"   # warm gold
E_LIGHT     = "EAF5EF"   # light mint tint
E_WHITE     = "FFFFFF"
E_GREEN     = "17A363"
E_RED       = "C0392B"
E_YELLOW    = "F39C12"

# ─────────────────────────────────────────────────────────────────────────────
# HELPER BUILDERS
# ─────────────────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border(color="999999"):
    t = Side(style="medium", color=color)
    s = Side(style="thin",   color=color)
    return Border(left=t, right=t, top=t, bottom=t)

def write(ws, row, col, value, fill_=None, font_=None, align_=None,
          border_=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill_:        cell.fill        = fill_
    if font_:        cell.font        = font_
    if align_:       cell.alignment   = align_
    if border_:      cell.border      = border_
    if number_format: cell.number_format = number_format
    return cell

def merge_write(ws, r1, c1, r2, c2, value, fill_=None, font_=None,
                align_=None, border_=None):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if fill_:   cell.fill      = fill_
    if font_:   cell.font      = font_
    if align_:  cell.alignment = align_
    if border_: cell.border    = border_
    return cell

def style_range(ws, r1, c1, r2, c2, fill_=None, font_=None,
                align_=None, border_=None, number_format=None):
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            if fill_:        cell.fill        = fill_
            if font_:        cell.font        = font_
            if align_:       cell.alignment   = align_
            if border_:      cell.border      = border_
            if number_format: cell.number_format = number_format

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
NEXT6  = []
now = datetime.date.today()
for i in range(6):
    month_idx = (now.month - 1 + i) % 12
    year      = now.year + ((now.month - 1 + i) // 12)
    NEXT6.append(f"{MONTHS[month_idx]} {year}")

PKR_FMT  = '#,##0.00 "PKR"'
USD_FMT  = '#,##0.00 "USD"'
PCT_FMT  = '0.00%'
NUM_FMT  = '#,##0.00'

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 – CRYPTO HOLDING & TRADING (SHARIA-COMPLIANT)
# ═════════════════════════════════════════════════════════════════════════════
def build_crypto_sheet(wb):
    ws = wb.create_sheet("Crypto Holdings & Trading")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B6"

    # column widths
    col_widths = [28, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18

    # ── MAIN TITLE ──────────────────────────────────────────────────────────
    merge_write(ws, 1, 1, 1, 15,
                "🪙  CRYPTO HOLDINGS & TRADING  |  SHARIA-COMPLIANT  |  STABLE COINS ONLY",
                fill_=fill(C_DARK),
                font_=font(bold=True, color=C_ACCENT, size=16),
                align_=align("center"))

    merge_write(ws, 2, 1, 2, 15,
                f"Business Intelligence Dashboard  |  Prepared: {datetime.date.today().strftime('%d %B %Y')}",
                fill_=fill(C_MID),
                font_=font(color=C_WHITE, size=10, italic=True),
                align_=align("center"))

    merge_write(ws, 3, 1, 3, 15, "",
                fill_=fill(C_ACCENT))

    # ════════════════════════════════════════════════════════════════════════
    # SECTION A – STABLE-COIN PORTFOLIO TRACKER
    # ════════════════════════════════════════════════════════════════════════
    r = 5
    merge_write(ws, r, 1, r, 15,
                "SECTION A  ─  STABLE-COIN PORTFOLIO TRACKER",
                fill_=fill(C_MID),
                font_=font(bold=True, color=C_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    hdr_cols_a = [
        "Coin / Wallet", "Exchange / Platform", "Opening Balance (USD)",
        "Purchases (USD)", "Sales / Withdrawals (USD)", "Fees Paid (USD)",
        "Profit/Gain (USD)", "Closing Balance (USD)", "PKR Equivalent",
        "Exchange Rate (PKR/USD)", "% of Portfolio", "Notes"
    ]
    # span 15 cols — merge last few for notes
    for ci, h in enumerate(hdr_cols_a, 1):
        write(ws, r, ci, h,
              fill_=fill(C_ACCENT),
              font_=font(bold=True, color=C_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "B8860B"))
    ws.row_dimensions[r].height = 32

    coins = ["USDT (Tether)", "USDC (USD Coin)", "BUSD / TUSD",
             "DAI", "Other Stable Coin 1", "Other Stable Coin 2"]
    for ci_row, coin in enumerate(coins):
        r += 1
        bg = fill(C_LIGHT) if ci_row % 2 == 0 else fill(C_WHITE)
        write(ws, r, 1, coin,  fill_=bg, font_=font(bold=True), border_=border())
        write(ws, r, 2, "",    fill_=bg, border_=border())    # platform
        write(ws, r, 3, 0.00,  fill_=bg, number_format=USD_FMT, border_=border())
        write(ws, r, 4, 0.00,  fill_=bg, number_format=USD_FMT, border_=border())
        write(ws, r, 5, 0.00,  fill_=bg, number_format=USD_FMT, border_=border())
        write(ws, r, 6, 0.00,  fill_=bg, number_format=USD_FMT, border_=border())
        # Gain = Closing - Opening - Purchases + Sales - Fees  (placeholder formula)
        c_open, c_buy, c_sell, c_fee = get_column_letter(3), get_column_letter(4), get_column_letter(5), get_column_letter(6)
        ws.cell(r, 7).value  = f"={get_column_letter(8)}{r}-{c_open}{r}-{c_buy}{r}+{c_sell}{r}-{c_fee}{r}"
        ws.cell(r, 7).fill   = bg
        ws.cell(r, 7).number_format = USD_FMT
        ws.cell(r, 7).border = border()
        write(ws, r, 8, 0.00,  fill_=bg, number_format=USD_FMT, border_=border())  # closing
        ws.cell(r, 9).value  = f"={get_column_letter(8)}{r}*{get_column_letter(10)}{r}"
        ws.cell(r, 9).fill   = bg
        ws.cell(r, 9).number_format = PKR_FMT
        ws.cell(r, 9).border = border()
        write(ws, r, 10, 278.00, fill_=bg, number_format=NUM_FMT, border_=border())  # exchange rate
        write(ws, r, 11, None,  fill_=bg, number_format=PCT_FMT, border_=border())   # % portfolio
        write(ws, r, 12, "",   fill_=bg, border_=border())

    # TOTAL ROW
    r += 1
    merge_write(ws, r, 1, r, 2, "TOTAL PORTFOLIO",
                fill_=fill(C_DARK), font_=font(bold=True, color=C_ACCENT),
                align_=align("center"), border_=border("medium", C_ACCENT))
    start_data = r - len(coins)
    for col in [3, 4, 5, 6, 8]:
        cl = get_column_letter(col)
        ws.cell(r, col).value         = f"=SUM({cl}{start_data}:{cl}{r-1})"
        ws.cell(r, col).fill          = fill(C_DARK)
        ws.cell(r, col).font          = font(bold=True, color=C_ACCENT)
        ws.cell(r, col).number_format = USD_FMT
        ws.cell(r, col).border        = border("medium", C_ACCENT)
    for col in [7]:
        cl = get_column_letter(col)
        ws.cell(r, col).value         = f"=SUM({cl}{start_data}:{cl}{r-1})"
        ws.cell(r, col).fill          = fill(C_DARK)
        ws.cell(r, col).font          = font(bold=True, color=C_ACCENT)
        ws.cell(r, col).number_format = USD_FMT
        ws.cell(r, col).border        = border("medium", C_ACCENT)
    for col in [9]:
        cl = get_column_letter(col)
        ws.cell(r, col).value         = f"=SUM({cl}{start_data}:{cl}{r-1})"
        ws.cell(r, col).fill          = fill(C_DARK)
        ws.cell(r, col).font          = font(bold=True, color=C_ACCENT)
        ws.cell(r, col).number_format = PKR_FMT
        ws.cell(r, col).border        = border("medium", C_ACCENT)
    for col in [10, 11, 12]:
        ws.cell(r, col).fill   = fill(C_DARK)
        ws.cell(r, col).border = border("medium", C_ACCENT)

    # ════════════════════════════════════════════════════════════════════════
    # SECTION B – MONTHLY EXPENSE TRACKER
    # ════════════════════════════════════════════════════════════════════════
    r += 2
    merge_write(ws, r, 1, r, 15,
                "SECTION B  ─  MONTHLY EXPENSE TRACKER",
                fill_=fill(C_MID),
                font_=font(bold=True, color=C_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    exp_headers = ["Expense Category", "Description"] + MONTHS[:12] + ["ANNUAL TOTAL"]
    for ci, h in enumerate(exp_headers, 1):
        write(ws, r, ci, h,
              fill_=fill(C_ACCENT),
              font_=font(bold=True, color=C_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "B8860B"))
    ws.row_dimensions[r].height = 28

    crypto_expenses = [
        ("Platform Fees",         "Exchange / withdrawal fees"),
        ("Blockchain Gas Fees",   "On-chain transaction costs"),
        ("Software & Tools",      "Portfolio trackers, analytics"),
        ("Regulatory / Legal",    "Compliance, KYC costs"),
        ("Banking Charges",       "Wire transfers, bank fees"),
        ("Office / Admin",        "Stationery, printing etc."),
        ("Zakat Provision",       "2.5% of nisab-eligible holdings"),
        ("Miscellaneous",         "Other unanticipated costs"),
    ]

    exp_start_r = r + 1
    for ei, (cat, desc) in enumerate(crypto_expenses):
        r += 1
        bg = fill(C_LIGHT) if ei % 2 == 0 else fill(C_WHITE)
        write(ws, r, 1, cat,  fill_=bg, font_=font(bold=True), border_=border())
        write(ws, r, 2, desc, fill_=bg, font_=font(italic=True, size=9), border_=border())
        for mc in range(3, 15):
            write(ws, r, mc, 0.00, fill_=bg, number_format=PKR_FMT, border_=border())
        total_col = 15
        data_range = f"{get_column_letter(3)}{r}:{get_column_letter(14)}{r}"
        ws.cell(r, total_col).value         = f"=SUM({data_range})"
        ws.cell(r, total_col).fill          = fill(C_MID)
        ws.cell(r, total_col).font          = font(bold=True, color=C_WHITE)
        ws.cell(r, total_col).number_format = PKR_FMT
        ws.cell(r, total_col).border        = border()

    r += 1
    merge_write(ws, r, 1, r, 2, "MONTHLY TOTAL",
                fill_=fill(C_DARK), font_=font(bold=True, color=C_ACCENT),
                align_=align("center"), border_=border("medium", C_ACCENT))
    for mc in range(3, 16):
        cl = get_column_letter(mc)
        ws.cell(r, mc).value         = f"=SUM({cl}{exp_start_r}:{cl}{r-1})"
        ws.cell(r, mc).fill          = fill(C_DARK)
        ws.cell(r, mc).font          = font(bold=True, color=C_ACCENT)
        ws.cell(r, mc).number_format = PKR_FMT
        ws.cell(r, mc).border        = border("medium", C_ACCENT)

    # ════════════════════════════════════════════════════════════════════════
    # SECTION C – 6-MONTH BUSINESS PROJECTION
    # ════════════════════════════════════════════════════════════════════════
    r += 2
    merge_write(ws, r, 1, r, 15,
                f"SECTION C  ─  6-MONTH BUSINESS PROJECTION  ({NEXT6[0]} – {NEXT6[5]})",
                fill_=fill(C_MID),
                font_=font(bold=True, color=C_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    proj_hdr = ["Item", "Metric"] + NEXT6 + ["6-Mo TOTAL", "Trend / Notes"]
    for ci, h in enumerate(proj_hdr, 1):
        write(ws, r, ci, h,
              fill_=fill(C_ACCENT),
              font_=font(bold=True, color=C_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "B8860B"))
    ws.row_dimensions[r].height = 28

    proj_items = [
        ("Total Portfolio Value (USD)",  "End-of-month balance"),
        ("Total Portfolio Value (PKR)",  "At prevailing rate"),
        ("New Purchases (USD)",          "Capital deployed"),
        ("Realized Gains (USD)",         "Closed profitable positions"),
        ("Unrealized Gains (USD)",       "Marked-to-market gain"),
        ("Total Fees Incurred (USD)",    "All platform & chain fees"),
        ("Zakat Provision (PKR)",        "Accrued 2.5% on holdings"),
        ("Net Portfolio Growth (USD)",   "Gains minus fees"),
        ("USDT Allocation (%)",          "% of portfolio"),
        ("USDC Allocation (%)",          "% of portfolio"),
        ("Cash Reserved (PKR)",          "Fiat on hand / in bank"),
        ("Expected Expenses (PKR)",      "Projected monthly costs"),
    ]

    proj_start_r = r + 1
    for pi, (item, metric) in enumerate(proj_items):
        r += 1
        bg = fill(C_LIGHT) if pi % 2 == 0 else fill(C_WHITE)
        write(ws, r, 1, item,   fill_=bg, font_=font(bold=True), border_=border())
        write(ws, r, 2, metric, fill_=bg, font_=font(italic=True, size=9), border_=border())
        for mc in range(3, 9):
            fmt = USD_FMT if "USD" in item else (PKR_FMT if "PKR" in item else PCT_FMT if "%" in item else NUM_FMT)
            write(ws, r, mc, 0.00, fill_=bg, number_format=fmt, border_=border())
        total_col  = 9
        notes_col  = 10
        data_range = f"{get_column_letter(3)}{r}:{get_column_letter(8)}{r}"
        agg = "SUM" if "%" not in item else "AVERAGE"
        fmt = USD_FMT if "USD" in item else (PKR_FMT if "PKR" in item else PCT_FMT if "%" in item else NUM_FMT)
        ws.cell(r, total_col).value         = f"={agg}({data_range})"
        ws.cell(r, total_col).fill          = fill(C_MID)
        ws.cell(r, total_col).font          = font(bold=True, color=C_WHITE)
        ws.cell(r, total_col).number_format = fmt
        ws.cell(r, total_col).border        = border()
        write(ws, r, notes_col, "",  fill_=bg, border_=border())
        # empty remaining cols to 15
        for ec in range(11, 16):
            write(ws, r, ec, "", fill_=bg, border_=border())

    # ════════════════════════════════════════════════════════════════════════
    # SECTION D – COSTING PROJECTION
    # ════════════════════════════════════════════════════════════════════════
    r += 2
    merge_write(ws, r, 1, r, 15,
                "SECTION D  ─  COSTING PROJECTION  (Per-Transaction & Operational)",
                fill_=fill(C_MID),
                font_=font(bold=True, color=C_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    cost_hdr = ["Cost Item", "Basis / Driver", "Unit Cost (USD)", "Est. Volume/Mo",
                "Monthly Cost (USD)", "Monthly Cost (PKR)", "6-Mo Projected (PKR)", "Notes"]
    for ci, h in enumerate(cost_hdr, 1):
        write(ws, r, ci, h,
              fill_=fill(C_ACCENT),
              font_=font(bold=True, color=C_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "B8860B"))
    ws.row_dimensions[r].height = 28

    cost_items = [
        ("Exchange Transaction Fee",   "Per trade (avg 0.1%)",    0.001,   20),
        ("Blockchain Transfer Fee",    "Per on-chain transfer",   2.50,    15),
        ("Crypto Wallet Hosting",      "Monthly subscription",    5.00,     1),
        ("Portfolio Tracker Software", "Monthly SaaS",           10.00,     1),
        ("KYC / Compliance Service",   "Annual ÷ 12",            20.00,     1),
        ("Accounting Software",        "Monthly",                15.00,     1),
        ("Legal / Advisory",           "Quarterly estimate",     50.00, 0.33),
        ("Internet / VPN",             "Per-month shared cost",   5.00,     1),
        ("Zakat Provision (USD eq.)",  "2.5% ÷ 12 of holdings",  0.00,     1),
        ("Miscellaneous",              "Buffer 5% of above",      0.00,     1),
    ]

    cost_start = r + 1
    exch_rate_ref = "278"   # fallback constant; user should link to portfolio tracker
    for cni, (item, basis, unit, vol) in enumerate(cost_items):
        r += 1
        bg = fill(C_LIGHT) if cni % 2 == 0 else fill(C_WHITE)
        write(ws, r, 1, item,    fill_=bg, font_=font(bold=True),      border_=border())
        write(ws, r, 2, basis,   fill_=bg, font_=font(italic=True, size=9), border_=border())
        write(ws, r, 3, unit,    fill_=bg, number_format=USD_FMT, border_=border())
        write(ws, r, 4, vol,     fill_=bg, number_format=NUM_FMT, border_=border())
        # Monthly USD = unit * vol
        ws.cell(r, 5).value = f"=C{r}*D{r}"
        ws.cell(r, 5).fill  = bg
        ws.cell(r, 5).number_format = USD_FMT
        ws.cell(r, 5).border = border()
        # Monthly PKR
        ws.cell(r, 6).value = f"=E{r}*{exch_rate_ref}"
        ws.cell(r, 6).fill  = bg
        ws.cell(r, 6).number_format = PKR_FMT
        ws.cell(r, 6).border = border()
        # 6-month
        ws.cell(r, 7).value = f"=F{r}*6"
        ws.cell(r, 7).fill  = bg
        ws.cell(r, 7).number_format = PKR_FMT
        ws.cell(r, 7).border = border()
        write(ws, r, 8, "", fill_=bg, border_=border())
        for ec in range(9, 16):
            write(ws, r, ec, "", fill_=fill(C_WHITE), border_=None)

    r += 1
    merge_write(ws, r, 1, r, 2, "TOTAL MONTHLY COSTS",
                fill_=fill(C_DARK), font_=font(bold=True, color=C_ACCENT),
                align_=align("center"), border_=border("medium", C_ACCENT))
    for col in [3, 5, 6, 7]:
        cl = get_column_letter(col)
        ws.cell(r, col).value         = f"=SUM({cl}{cost_start}:{cl}{r-1})"
        ws.cell(r, col).fill          = fill(C_DARK)
        ws.cell(r, col).font          = font(bold=True, color=C_ACCENT)
        ws.cell(r, col).number_format = USD_FMT if col in [3,5] else PKR_FMT
        ws.cell(r, col).border        = border("medium", C_ACCENT)
    ws.cell(r, 4).fill   = fill(C_DARK); ws.cell(r, 4).border = border("medium", C_ACCENT)
    ws.cell(r, 8).fill   = fill(C_DARK); ws.cell(r, 8).border = border("medium", C_ACCENT)

    # ════════════════════════════════════════════════════════════════════════
    # SECTION E – RESOURCE REQUISITION PROJECTION
    # ════════════════════════════════════════════════════════════════════════
    r += 2
    merge_write(ws, r, 1, r, 15,
                "SECTION E  ─  RESOURCE REQUISITION PROJECTION  (Next 6 Months)",
                fill_=fill(C_MID),
                font_=font(bold=True, color=C_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    res_hdr = ["Resource / Item", "Category", "Quantity", "Unit",
               "Unit Cost (PKR)", "Total Cost (PKR)", "Required By",
               "Priority", "Vendor / Source", "Status", "Notes"]
    for ci, h in enumerate(res_hdr, 1):
        write(ws, r, ci, h,
              fill_=fill(C_ACCENT),
              font_=font(bold=True, color=C_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "B8860B"))
    ws.row_dimensions[r].height = 28

    resources = [
        ("High-speed Internet Upgrade", "Infrastructure",  1, "Month",   5000, NEXT6[0], "High",    "ISP Provider",      "Planned"),
        ("Hardware Wallet (Ledger)",    "Security",        2, "Unit",    25000, NEXT6[0], "Critical","Trusted E-commerce","Planned"),
        ("VPN Subscription",            "Security",        1, "Year",    12000, NEXT6[0], "High",    "ExpressVPN/similar","Planned"),
        ("Portfolio Analytics Tool",    "Software",        1, "Year",    18000, NEXT6[1], "Medium",  "Online SaaS",       "Evaluating"),
        ("Accounting Software License", "Software",        1, "Year",    24000, NEXT6[1], "High",    "QuickBooks/Zoho",   "Planned"),
        ("Dedicated Workstation",       "Hardware",        1, "Unit",   120000, NEXT6[2], "Medium",  "Local Vendor",      "Pending"),
        ("UPS / Power Backup",          "Hardware",        1, "Unit",    35000, NEXT6[2], "Medium",  "Local Vendor",      "Pending"),
        ("Legal/Compliance Consultant", "Professional",    1, "Quarter", 50000, NEXT6[3], "High",    "Legal Firm",        "Sourcing"),
        ("External Backup Drive (2TB)", "Storage",         2, "Unit",    15000, NEXT6[4], "Medium",  "Trusted E-commerce","Pending"),
        ("Training / Courses",          "Development",     2, "Course",   8000, NEXT6[5], "Low",     "Udemy / Institute", "Planned"),
    ]

    for ri, res in enumerate(resources):
        r += 1
        bg = fill(C_LIGHT) if ri % 2 == 0 else fill(C_WHITE)
        item, cat, qty, unit, unit_cost, req_by, priority, vendor, status = res
        write(ws, r, 1,  item,      fill_=bg, font_=font(bold=True), border_=border())
        write(ws, r, 2,  cat,       fill_=bg, border_=border())
        write(ws, r, 3,  qty,       fill_=bg, number_format=NUM_FMT, border_=border())
        write(ws, r, 4,  unit,      fill_=bg, border_=border())
        write(ws, r, 5,  unit_cost, fill_=bg, number_format=PKR_FMT, border_=border())
        ws.cell(r, 6).value         = f"=C{r}*E{r}"
        ws.cell(r, 6).fill          = bg
        ws.cell(r, 6).number_format = PKR_FMT
        ws.cell(r, 6).border        = border()
        write(ws, r, 7,  req_by,    fill_=bg, border_=border())
        # priority colour
        p_color = {"Critical": C_RED, "High": C_YELLOW, "Medium": "4472C4", "Low": C_GREEN}.get(priority, C_WHITE)
        write(ws, r, 8, priority, fill_=fill(p_color),
              font_=font(bold=True, color=C_WHITE if priority != "Medium" else C_WHITE),
              align_=align("center"), border_=border())
        write(ws, r, 9,  vendor,    fill_=bg, border_=border())
        write(ws, r, 10, status,    fill_=bg, border_=border())
        write(ws, r, 11, "",        fill_=bg, border_=border())
        for ec in range(12, 16):
            write(ws, r, ec, "", fill_=fill(C_WHITE))

    r += 1
    merge_write(ws, r, 1, r, 5, "TOTAL RESOURCE BUDGET (PKR)",
                fill_=fill(C_DARK), font_=font(bold=True, color=C_ACCENT),
                align_=align("center"), border_=border("medium", C_ACCENT))
    res_start = r - len(resources)
    ws.cell(r, 6).value         = f"=SUM(F{res_start}:F{r-1})"
    ws.cell(r, 6).fill          = fill(C_DARK)
    ws.cell(r, 6).font          = font(bold=True, color=C_ACCENT)
    ws.cell(r, 6).number_format = PKR_FMT
    ws.cell(r, 6).border        = border("medium", C_ACCENT)
    for ec in range(7, 16):
        ws.cell(r, ec).fill   = fill(C_DARK)
        ws.cell(r, ec).border = border("medium", C_ACCENT)

    # footer
    r += 2
    merge_write(ws, r, 1, r, 15,
                "⚠  SHARIA NOTE: This sheet tracks ONLY spot transactions in stable coins. No futures, derivatives, leverage, margin, or interest-bearing instruments are included or permitted.",
                fill_=fill("FFF3CD"),
                font_=font(italic=True, color="7D4E00", size=9),
                align_=align("center", wrap=True))
    ws.row_dimensions[r].height = 32

    return ws


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 – EQAAZ ISLAMIC SCHOOLING SYSTEM
# ═════════════════════════════════════════════════════════════════════════════
def build_eqaaz_sheet(wb):
    ws = wb.create_sheet("Eqaaz Islamic School")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B6"

    col_widths = [28, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18

    # ── MAIN TITLE ──────────────────────────────────────────────────────────
    merge_write(ws, 1, 1, 1, 15,
                "🕌  EQAAZ ISLAMIC SCHOOLING SYSTEM  |  BUSINESS MANAGEMENT DASHBOARD",
                fill_=fill(E_DARK),
                font_=font(bold=True, color=E_ACCENT, size=16),
                align_=align("center"))

    merge_write(ws, 2, 1, 2, 15,
                f"Rental Property Setup  ·  Medium Scale  |  Prepared: {datetime.date.today().strftime('%d %B %Y')}",
                fill_=fill(E_MID),
                font_=font(color=E_WHITE, size=10, italic=True),
                align_=align("center"))

    merge_write(ws, 3, 1, 3, 15, "",
                fill_=fill(E_ACCENT))

    # ════════════════════════════════════════════════════════════════════════
    # SECTION A – MONTHLY INCOME TRACKER
    # ════════════════════════════════════════════════════════════════════════
    r = 5
    merge_write(ws, r, 1, r, 15,
                "SECTION A  ─  MONTHLY INCOME TRACKER",
                fill_=fill(E_MID),
                font_=font(bold=True, color=E_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    inc_hdr = ["Income Source", "Description"] + MONTHS[:12] + ["ANNUAL TOTAL"]
    for ci, h in enumerate(inc_hdr, 1):
        write(ws, r, ci, h,
              fill_=fill(E_ACCENT),
              font_=font(bold=True, color=E_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "8B7332"))
    ws.row_dimensions[r].height = 28

    income_items = [
        ("Monthly Tuition Fees",    "Per-student monthly fee × enrolment"),
        ("Admission / Registration","One-time per new student"),
        ("Quran Hifz Program Fees", "Dedicated hifz class fees"),
        ("Nazra / Basic Quran",     "Separate nazra batch"),
        ("Islamic Studies Fees",    "Optional subject premium"),
        ("Exam / Assessment Fees",  "Per assessment cycle"),
        ("Stationery / Kit Sales",  "Books, uniform packages"),
        ("Donation / Zakat Funds",  "Zakat / Sadaqah received"),
        ("Government / NGO Grant",  "If applicable"),
        ("Other Income",            "Miscellaneous"),
    ]

    inc_start = r + 1
    for ii, (src, desc) in enumerate(income_items):
        r += 1
        bg = fill(E_LIGHT) if ii % 2 == 0 else fill(E_WHITE)
        write(ws, r, 1, src,  fill_=bg, font_=font(bold=True), border_=border())
        write(ws, r, 2, desc, fill_=bg, font_=font(italic=True, size=9), border_=border())
        for mc in range(3, 15):
            write(ws, r, mc, 0.00, fill_=bg, number_format=PKR_FMT, border_=border())
        total_col = 15
        ws.cell(r, total_col).value         = f"=SUM({get_column_letter(3)}{r}:{get_column_letter(14)}{r})"
        ws.cell(r, total_col).fill          = fill(E_MID)
        ws.cell(r, total_col).font          = font(bold=True, color=E_WHITE)
        ws.cell(r, total_col).number_format = PKR_FMT
        ws.cell(r, total_col).border        = border()

    r += 1
    merge_write(ws, r, 1, r, 2, "TOTAL MONTHLY INCOME",
                fill_=fill(E_DARK), font_=font(bold=True, color=E_ACCENT),
                align_=align("center"), border_=border("medium", E_ACCENT))
    for mc in range(3, 16):
        cl = get_column_letter(mc)
        ws.cell(r, mc).value         = f"=SUM({cl}{inc_start}:{cl}{r-1})"
        ws.cell(r, mc).fill          = fill(E_DARK)
        ws.cell(r, mc).font          = font(bold=True, color=E_ACCENT)
        ws.cell(r, mc).number_format = PKR_FMT
        ws.cell(r, mc).border        = border("medium", E_ACCENT)

    total_income_row = r

    # ════════════════════════════════════════════════════════════════════════
    # SECTION B – MONTHLY EXPENSE TRACKER
    # ════════════════════════════════════════════════════════════════════════
    r += 2
    merge_write(ws, r, 1, r, 15,
                "SECTION B  ─  MONTHLY EXPENSE TRACKER",
                fill_=fill(E_MID),
                font_=font(bold=True, color=E_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    exp_hdr = ["Expense Category", "Description"] + MONTHS[:12] + ["ANNUAL TOTAL"]
    for ci, h in enumerate(exp_hdr, 1):
        write(ws, r, ci, h,
              fill_=fill(E_ACCENT),
              font_=font(bold=True, color=E_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "8B7332"))
    ws.row_dimensions[r].height = 28

    eqaaz_expenses = [
        ("Property Rent",             "Monthly rental (full premises)"),
        ("Electricity Bill",          "WAPDA / LESCO charges"),
        ("Gas / Heating",             "Sui Gas or cylinder"),
        ("Water / WASA",              "Monthly water charges"),
        ("Internet & Telephony",      "Broadband + phone bills"),
        ("Teaching Staff Salaries",   "Quran + academics teachers"),
        ("Admin / Support Staff",     "Receptionist, peon, guard"),
        ("Driver / Transport",        "School van / fuel if applicable"),
        ("Cleaning & Maintenance",    "Janitor + repairs"),
        ("Teaching Materials",        "Books, worksheets, stationery"),
        ("IT Equipment Maintenance",  "Projector, PC, printer repair"),
        ("Prayer Hall / Mosque Upkeep","Quran stands, prayer mats, oud"),
        ("First Aid / Health",        "Medicine box, safety items"),
        ("Marketing & Admission",     "Banners, social media, flyers"),
        ("Zakat / Charity Disbursed", "From donations received"),
        ("Miscellaneous & Contingency","Unexpected expenses buffer"),
    ]

    exp_start_e = r + 1
    for ei, (cat, desc) in enumerate(eqaaz_expenses):
        r += 1
        bg = fill(E_LIGHT) if ei % 2 == 0 else fill(E_WHITE)
        write(ws, r, 1, cat,  fill_=bg, font_=font(bold=True), border_=border())
        write(ws, r, 2, desc, fill_=bg, font_=font(italic=True, size=9), border_=border())
        for mc in range(3, 15):
            write(ws, r, mc, 0.00, fill_=bg, number_format=PKR_FMT, border_=border())
        ws.cell(r, 15).value         = f"=SUM({get_column_letter(3)}{r}:{get_column_letter(14)}{r})"
        ws.cell(r, 15).fill          = fill(E_MID)
        ws.cell(r, 15).font          = font(bold=True, color=E_WHITE)
        ws.cell(r, 15).number_format = PKR_FMT
        ws.cell(r, 15).border        = border()

    r += 1
    merge_write(ws, r, 1, r, 2, "TOTAL MONTHLY EXPENSES",
                fill_=fill(E_DARK), font_=font(bold=True, color=E_ACCENT),
                align_=align("center"), border_=border("medium", E_ACCENT))
    for mc in range(3, 16):
        cl = get_column_letter(mc)
        ws.cell(r, mc).value         = f"=SUM({cl}{exp_start_e}:{cl}{r-1})"
        ws.cell(r, mc).fill          = fill(E_DARK)
        ws.cell(r, mc).font          = font(bold=True, color=E_ACCENT)
        ws.cell(r, mc).number_format = PKR_FMT
        ws.cell(r, mc).border        = border("medium", E_ACCENT)

    total_exp_row = r

    # NET SURPLUS / DEFICIT
    r += 1
    merge_write(ws, r, 1, r, 2, "NET SURPLUS / (DEFICIT)",
                fill_=fill("1A6635"), font_=font(bold=True, color=E_WHITE),
                align_=align("center"), border_=border("medium", E_ACCENT))
    for mc in range(3, 16):
        cl = get_column_letter(mc)
        ws.cell(r, mc).value         = f"={cl}{total_income_row}-{cl}{total_exp_row}"
        ws.cell(r, mc).fill          = fill("1A6635")
        ws.cell(r, mc).font          = font(bold=True, color=E_WHITE)
        ws.cell(r, mc).number_format = PKR_FMT
        ws.cell(r, mc).border        = border("medium", E_ACCENT)

    # ════════════════════════════════════════════════════════════════════════
    # SECTION C – 6-MONTH BUSINESS PROJECTION
    # ════════════════════════════════════════════════════════════════════════
    r += 2
    merge_write(ws, r, 1, r, 15,
                f"SECTION C  ─  6-MONTH BUSINESS PROJECTION  ({NEXT6[0]} – {NEXT6[5]})",
                fill_=fill(E_MID),
                font_=font(bold=True, color=E_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    proj_hdr = ["Projection Item", "Assumption / Basis"] + NEXT6 + ["6-Mo TOTAL", "Notes"]
    for ci, h in enumerate(proj_hdr, 1):
        write(ws, r, ci, h,
              fill_=fill(E_ACCENT),
              font_=font(bold=True, color=E_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "8B7332"))
    ws.row_dimensions[r].height = 28

    eqaaz_proj = [
        ("Expected Enrolment (Students)",   "No. of enrolled students"),
        ("Avg Monthly Fee per Student (PKR)","Blended average"),
        ("Projected Tuition Revenue (PKR)",  "Enrolment × Avg Fee"),
        ("Other Income (PKR)",               "Admissions + grants etc."),
        ("Total Projected Income (PKR)",     "Sum of all income"),
        ("Staff Salaries (PKR)",             "Teaching + admin"),
        ("Rental Cost (PKR)",                "Fixed monthly rent"),
        ("Utilities (PKR)",                  "Electricity + gas + water"),
        ("Operations Budget (PKR)",          "Materials + misc"),
        ("Total Projected Expenses (PKR)",   "Sum of all expenses"),
        ("Projected Surplus / Deficit (PKR)","Income – Expenses"),
        ("Break-even Enrolment",             "Expenses ÷ Avg Fee"),
        ("Capacity Utilisation (%)",         "Enrolled ÷ Max capacity"),
    ]

    for pi, (item, basis) in enumerate(eqaaz_proj):
        r += 1
        bg = fill(E_LIGHT) if pi % 2 == 0 else fill(E_WHITE)
        write(ws, r, 1, item,  fill_=bg, font_=font(bold=True), border_=border())
        write(ws, r, 2, basis, fill_=bg, font_=font(italic=True, size=9), border_=border())
        for mc in range(3, 9):
            fmt = NUM_FMT if "Students" in item or "Enrolment" in item or "%" in item else PKR_FMT
            write(ws, r, mc, 0.00, fill_=bg, number_format=fmt, border_=border())
        data_range = f"{get_column_letter(3)}{r}:{get_column_letter(8)}{r}"
        agg = "AVERAGE" if "%" in item else "SUM"
        ws.cell(r, 9).value         = f"={agg}({data_range})"
        ws.cell(r, 9).fill          = fill(E_MID)
        ws.cell(r, 9).font          = font(bold=True, color=E_WHITE)
        ws.cell(r, 9).number_format = PKR_FMT if "PKR" in item else NUM_FMT
        ws.cell(r, 9).border        = border()
        write(ws, r, 10, "", fill_=bg, border_=border())
        for ec in range(11, 16):
            write(ws, r, ec, "", fill_=fill(E_WHITE))

    # ════════════════════════════════════════════════════════════════════════
    # SECTION D – COSTING PROJECTION
    # ════════════════════════════════════════════════════════════════════════
    r += 2
    merge_write(ws, r, 1, r, 15,
                "SECTION D  ─  COSTING PROJECTION  (Per-Student & Operational)",
                fill_=fill(E_MID),
                font_=font(bold=True, color=E_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    cost_hdr2 = ["Cost Item", "Cost Driver / Basis", "Unit Cost (PKR)", "Est. Units/Mo",
                 "Monthly Cost (PKR)", "Annual Cost (PKR)", "Cost per Student (PKR)", "Notes"]
    for ci, h in enumerate(cost_hdr2, 1):
        write(ws, r, ci, h,
              fill_=fill(E_ACCENT),
              font_=font(bold=True, color=E_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "8B7332"))
    ws.row_dimensions[r].height = 28

    cost_items2 = [
        ("Teaching Staff – Quran",    "Per teacher × headcount",   35000, 3),
        ("Teaching Staff – Academic", "Per teacher × headcount",   30000, 2),
        ("Admin / Non-Teaching",      "Receptionist + peon etc.",  20000, 2),
        ("Rental – Main Building",    "Full premises monthly",     80000, 1),
        ("Electricity + Gas",         "Estimated utility bundle",  15000, 1),
        ("Internet & Telephony",       "Broadband + SIM",           3000, 1),
        ("Books & Stationery",        "Per-student per month",      1500, 1),  # × enrolment
        ("Maintenance & Repair",      "Monthly provision",          8000, 1),
        ("Marketing / Admissions",    "Banners + digital + print",  5000, 1),
        ("First Aid & Safety",        "Monthly provision",          2000, 1),
        ("Prayer Hall Supplies",      "Quran stands, mats, oud",    3000, 1),
        ("Miscellaneous Buffer",      "5% contingency",             0,    1),
    ]

    eqaaz_cost_start = r + 1
    for cni, (item, basis, unit, vol) in enumerate(cost_items2):
        r += 1
        bg = fill(E_LIGHT) if cni % 2 == 0 else fill(E_WHITE)
        write(ws, r, 1, item,    fill_=bg, font_=font(bold=True),      border_=border())
        write(ws, r, 2, basis,   fill_=bg, font_=font(italic=True, size=9), border_=border())
        write(ws, r, 3, unit,    fill_=bg, number_format=PKR_FMT, border_=border())
        write(ws, r, 4, vol,     fill_=bg, number_format=NUM_FMT, border_=border())
        ws.cell(r, 5).value         = f"=C{r}*D{r}"
        ws.cell(r, 5).fill          = bg
        ws.cell(r, 5).number_format = PKR_FMT
        ws.cell(r, 5).border        = border()
        ws.cell(r, 6).value         = f"=E{r}*12"
        ws.cell(r, 6).fill          = bg
        ws.cell(r, 6).number_format = PKR_FMT
        ws.cell(r, 6).border        = border()
        # cost per student (assuming ~60 as placeholder, user can change)
        ws.cell(r, 7).value         = f"=E{r}/60"
        ws.cell(r, 7).fill          = bg
        ws.cell(r, 7).number_format = PKR_FMT
        ws.cell(r, 7).border        = border()
        write(ws, r, 8, "", fill_=bg, border_=border())
        for ec in range(9, 16):
            write(ws, r, ec, "", fill_=fill(E_WHITE))

    r += 1
    merge_write(ws, r, 1, r, 2, "TOTAL MONTHLY OPERATING COST",
                fill_=fill(E_DARK), font_=font(bold=True, color=E_ACCENT),
                align_=align("center"), border_=border("medium", E_ACCENT))
    for col in [5, 6]:
        cl = get_column_letter(col)
        ws.cell(r, col).value         = f"=SUM({cl}{eqaaz_cost_start}:{cl}{r-1})"
        ws.cell(r, col).fill          = fill(E_DARK)
        ws.cell(r, col).font          = font(bold=True, color=E_ACCENT)
        ws.cell(r, col).number_format = PKR_FMT
        ws.cell(r, col).border        = border("medium", E_ACCENT)
    for col in [3, 4, 7, 8]:
        ws.cell(r, col).fill   = fill(E_DARK)
        ws.cell(r, col).border = border("medium", E_ACCENT)

    # ════════════════════════════════════════════════════════════════════════
    # SECTION E – RESOURCE REQUISITION PROJECTION
    # ════════════════════════════════════════════════════════════════════════
    r += 2
    merge_write(ws, r, 1, r, 15,
                "SECTION E  ─  RESOURCE REQUISITION PROJECTION  (Next 6 Months)",
                fill_=fill(E_MID),
                font_=font(bold=True, color=E_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    res_hdr2 = ["Resource / Item", "Category", "Qty", "Unit",
                "Unit Cost (PKR)", "Total Cost (PKR)", "Required By",
                "Priority", "Vendor / Source", "Status", "Notes"]
    for ci, h in enumerate(res_hdr2, 1):
        write(ws, r, ci, h,
              fill_=fill(E_ACCENT),
              font_=font(bold=True, color=E_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "8B7332"))
    ws.row_dimensions[r].height = 28

    eqaaz_resources = [
        ("CCTV Security System",       "Infrastructure", 6,  "Camera", 12000, NEXT6[0], "High",    "Local Vendor",       "Planned"),
        ("Whiteboards (Large)",        "Furniture",      8,  "Unit",    4500, NEXT6[0], "High",    "Stationers",         "Planned"),
        ("Student Chairs & Desks",     "Furniture",      30, "Set",     3500, NEXT6[0], "Critical","Local Manufacturer", "Ordered"),
        ("Quran Stands (Rehal)",       "Prayer Supplies",20, "Unit",     800, NEXT6[0], "Medium",  "Islamic Bookstore",  "Planned"),
        ("Prayer Mats",                "Prayer Supplies",25, "Mat",     1200, NEXT6[1], "Medium",  "Islamic Bookstore",  "Planned"),
        ("Projector & Screen",         "AV Equipment",   2,  "Set",    45000, NEXT6[1], "High",    "Electronics Shop",   "Pending"),
        ("Printer (A4 Laser)",         "IT Equipment",   2,  "Unit",    35000, NEXT6[1], "High",    "IT Vendor",          "Pending"),
        ("First Aid Kit",              "Health & Safety",3,  "Kit",     3000, NEXT6[2], "High",    "Pharmacy",           "Planned"),
        ("Fire Extinguishers",         "Safety",         4,  "Unit",    5500, NEXT6[2], "Critical","Safety Supplier",    "Planned"),
        ("Air Coolers / AC Units",     "Climate Control",4,  "Unit",   55000, NEXT6[2], "High",    "Electronics Shop",   "Evaluating"),
        ("School Signboard & Branding","Marketing",       1,  "Set",    25000, NEXT6[3], "Medium",  "Signage Printer",    "Planned"),
        ("Library Books (Islamic)",    "Library",        100,"Book",    1500, NEXT6[3], "Medium",  "Islamic Publishers", "Sourcing"),
        ("Computer Lab (Basic)",       "IT Equipment",   5,  "Unit",   95000, NEXT6[4], "Low",     "IT Vendor",          "Evaluating"),
        ("Generator / UPS",            "Power Backup",   1,  "Unit",   85000, NEXT6[4], "High",    "Local Vendor",       "Pending"),
        ("Sports / Activity Equipment","Co-Curricular",  1,  "Set",    20000, NEXT6[5], "Low",     "Sports Shop",        "Planned"),
    ]

    eqaaz_res_start = r + 1
    for ri, res in enumerate(eqaaz_resources):
        r += 1
        bg = fill(E_LIGHT) if ri % 2 == 0 else fill(E_WHITE)
        item, cat, qty, unit, unit_cost, req_by, priority, vendor, status = res
        write(ws, r, 1, item,      fill_=bg, font_=font(bold=True), border_=border())
        write(ws, r, 2, cat,       fill_=bg, border_=border())
        write(ws, r, 3, qty,       fill_=bg, number_format=NUM_FMT, border_=border())
        write(ws, r, 4, unit,      fill_=bg, border_=border())
        write(ws, r, 5, unit_cost, fill_=bg, number_format=PKR_FMT, border_=border())
        ws.cell(r, 6).value         = f"=C{r}*E{r}"
        ws.cell(r, 6).fill          = bg
        ws.cell(r, 6).number_format = PKR_FMT
        ws.cell(r, 6).border        = border()
        write(ws, r, 7, req_by,    fill_=bg, border_=border())
        p_color = {"Critical": E_RED, "High": E_YELLOW, "Medium": "0A7A50", "Low": E_GREEN}.get(priority, E_WHITE)
        write(ws, r, 8, priority, fill_=fill(p_color),
              font_=font(bold=True, color=E_WHITE),
              align_=align("center"), border_=border())
        write(ws, r, 9,  vendor,  fill_=bg, border_=border())
        write(ws, r, 10, status,  fill_=bg, border_=border())
        write(ws, r, 11, "",      fill_=bg, border_=border())
        for ec in range(12, 16):
            write(ws, r, ec, "", fill_=fill(E_WHITE))

    r += 1
    merge_write(ws, r, 1, r, 5, "TOTAL RESOURCE BUDGET (PKR)",
                fill_=fill(E_DARK), font_=font(bold=True, color=E_ACCENT),
                align_=align("center"), border_=border("medium", E_ACCENT))
    ws.cell(r, 6).value         = f"=SUM(F{eqaaz_res_start}:F{r-1})"
    ws.cell(r, 6).fill          = fill(E_DARK)
    ws.cell(r, 6).font          = font(bold=True, color=E_ACCENT)
    ws.cell(r, 6).number_format = PKR_FMT
    ws.cell(r, 6).border        = border("medium", E_ACCENT)
    for ec in range(7, 16):
        ws.cell(r, ec).fill   = fill(E_DARK)
        ws.cell(r, ec).border = border("medium", E_ACCENT)

    # ════════════════════════════════════════════════════════════════════════
    # SECTION F – STAFF ROSTER & SALARY REGISTER
    # ════════════════════════════════════════════════════════════════════════
    r += 2
    merge_write(ws, r, 1, r, 15,
                "SECTION F  ─  STAFF ROSTER & MONTHLY SALARY REGISTER",
                fill_=fill(E_MID),
                font_=font(bold=True, color=E_ACCENT, size=12),
                align_=align("left"))
    ws.row_dimensions[r].height = 24

    r += 1
    staff_hdr = ["Staff Name", "Designation", "Department", "Joining Date",
                 "Basic Salary (PKR)", "Allowances (PKR)", "Deductions (PKR)",
                 "Net Salary (PKR)",   "Payment Status", "Notes"]
    for ci, h in enumerate(staff_hdr, 1):
        write(ws, r, ci, h,
              fill_=fill(E_ACCENT),
              font_=font(bold=True, color=E_DARK, size=9),
              align_=align("center", wrap=True),
              border_=border("thin", "8B7332"))
    ws.row_dimensions[r].height = 28

    staff_templates = [
        ("Ustadh / Hafiz [Name]",  "Quran Teacher",     "Hifz Dept."),
        ("Ustadh / Hafiz [Name]",  "Quran Teacher",     "Nazra Dept."),
        ("Ustadh / Hafiz [Name]",  "Quran Teacher",     "Hifz Dept."),
        ("Teacher [Name]",         "Academic Teacher",  "General Studies"),
        ("Teacher [Name]",         "Academic Teacher",  "General Studies"),
        ("Admin [Name]",           "Administrator",     "Admin"),
        ("Receptionist [Name]",    "Receptionist",      "Admin"),
        ("Security Guard [Name]",  "Security",          "Support"),
        ("Peon / Cleaner [Name]",  "Support Staff",     "Support"),
    ]

    staff_start = r + 1
    for si, (name, desig, dept) in enumerate(staff_templates):
        r += 1
        bg = fill(E_LIGHT) if si % 2 == 0 else fill(E_WHITE)
        write(ws, r, 1,  name,  fill_=bg, font_=font(italic=True), border_=border())
        write(ws, r, 2,  desig, fill_=bg, border_=border())
        write(ws, r, 3,  dept,  fill_=bg, border_=border())
        write(ws, r, 4,  "",    fill_=bg, border_=border())   # joining date
        write(ws, r, 5,  0.00,  fill_=bg, number_format=PKR_FMT, border_=border())
        write(ws, r, 6,  0.00,  fill_=bg, number_format=PKR_FMT, border_=border())
        write(ws, r, 7,  0.00,  fill_=bg, number_format=PKR_FMT, border_=border())
        ws.cell(r, 8).value         = f"=E{r}+F{r}-G{r}"
        ws.cell(r, 8).fill          = bg
        ws.cell(r, 8).number_format = PKR_FMT
        ws.cell(r, 8).border        = border()
        write(ws, r, 9,  "Pending", fill_=fill(C_YELLOW),
              font_=font(bold=True, color="5D4000"), align_=align("center"), border_=border())
        write(ws, r, 10, "",  fill_=bg, border_=border())
        for ec in range(11, 16):
            write(ws, r, ec, "", fill_=fill(E_WHITE))

    r += 1
    merge_write(ws, r, 1, r, 4, "TOTAL MONTHLY PAYROLL",
                fill_=fill(E_DARK), font_=font(bold=True, color=E_ACCENT),
                align_=align("center"), border_=border("medium", E_ACCENT))
    for col in [5, 6, 7, 8]:
        cl = get_column_letter(col)
        ws.cell(r, col).value         = f"=SUM({cl}{staff_start}:{cl}{r-1})"
        ws.cell(r, col).fill          = fill(E_DARK)
        ws.cell(r, col).font          = font(bold=True, color=E_ACCENT)
        ws.cell(r, col).number_format = PKR_FMT
        ws.cell(r, col).border        = border("medium", E_ACCENT)
    for col in [9, 10]:
        ws.cell(r, col).fill   = fill(E_DARK)
        ws.cell(r, col).border = border("medium", E_ACCENT)

    # footer
    r += 2
    merge_write(ws, r, 1, r, 15,
                "بسم الله الرحمن الرحيم  ·  May Allah bless this institution and all who serve it.  ·  Eqaaz Islamic Schooling System",
                fill_=fill("EAF5EF"),
                font_=font(italic=True, color=E_DARK, size=10),
                align_=align("center", wrap=True))
    ws.row_dimensions[r].height = 28

    return ws


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
# Remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

build_crypto_sheet(wb)
build_eqaaz_sheet(wb)

output_path = r"d:\7ctech\eqaz system\Business_Dashboard.xlsx"
wb.save(output_path)
print(f"[OK] Workbook saved to: {output_path}")
